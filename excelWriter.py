import openpyxl
from tqdm import tqdm
from oplan import *

class WriterExcel:

    def __init__(self, path):
        self.path = path
        self.wb = openpyxl.Workbook()
        self.sheets = list()

    def writeData(self, sheet, row, column, data, color=None):
        sheet.cell(row=row, column=column).value = data
        sheet.cell(row=row, column=column).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        if color != None:
            sheet.cell(row=row, column=column).fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type='solid')
        self.wb.save(self.path)

    def writeConstraintes(self, teachers, schedule):
        sheet = self.wb.create_sheet(index = len(self.sheets) , title = "constraints")
        self.sheets.append(sheet)

        dictDays = {day: i+2 for i, day in enumerate(DAYS)}
        dictHours = {hour: i+2 for i, hour in enumerate(HOURS)}
        dictClasses = {classroom: i+2 for i, classroom in enumerate(CLASSES)}

        index = 0
        for teacher in teachers:
            for hour, column  in dictHours.items():
                self.writeData(sheet, 1 + index, column, hour)
            for day, row in dictDays.items():
                for classroom, row2 in dictClasses.items():
                    self.writeData(sheet, len(dictClasses)*(row-2) + row2 + index , 1, f"{day} / {classroom}")
            self.writeData(sheet, index+1, 1, teacher.id)
            for constraint in teacher.constraints:
                for c in range(len(dictClasses)):
                    self.writeData(sheet, len(dictClasses)*(dictDays[constraint[0]]-2) + c+2 + index, dictHours[constraint[1]], "X", "000000")
                    
            index += len(dictDays)*len(dictClasses)+2



    def writeSchedule(self, schedule):
        sheet = self.wb.create_sheet(index = len(self.sheets) , title = "schedule")
        self.sheets.append(sheet)
        self.writeData(sheet, 1, 1, schedule.getScore())

        

        dictDays = {day: i+2 for i, day in enumerate(DAYS)}
        dictHours = {hour: i+2 for i, hour in enumerate(HOURS)}
        dictClasses = {classroom: i+2 for i, classroom in enumerate(CLASSES)}
        for hour, column  in dictHours.items():
            self.writeData(sheet, 1, column, hour)
        for day, row in dictDays.items():
            for classroom, row2 in dictClasses.items():
                self.writeData(sheet, len(dictClasses)*(row-2) + row2, 1, f"{day} / {classroom}")

        for slot in schedule.schedule:
            period = slot.period
            for index, defense in enumerate(slot.defenses):
                self.writeData(sheet, len(dictClasses)*(dictDays[period.day]-2) + index+2, dictHours[period.hour], f"{defense[1].id}  / {defense[0].getStudent().id} / {defense[0].getTutor().id} / {defense[0].getApm().id}")



if __name__ == "__main__":
    sm = ScheduleManager()

    # student1 = Person("Allan", "Des Courtils", 1)
    # student2 = Person("Thibault", "Thomas", 2)
    # student3 = Person("Emilie", "Vey", 3)
    # student4 = Person("Nathan", "Peyronnet", 4)
    # student5 = Person("Benoit", "Kezel", 5)

    # teacher1 = Teacher("Moahammed", "Haddad", 6, [slots[random.randint(0, len(slots)-1)] for i in range(random.randint(0, len(slots)-1))])
    # teacher2 = Teacher("Bastien", "Jerome", 7, [slots[random.randint(0, len(slots)-1)] for i in range(random.randint(0, len(slots)-1))])
    # teacher3 = Teacher("Valerie", "James", 8, [slots[random.randint(0, len(slots)-1)] for i in range(random.randint(0, len(slots)-1))])
    # teacher4 = Teacher("Florence", "Perraud", 9, [slots[random.randint(0, len(slots)-1)] for i in range(random.randint(0, len(slots)-1))])
    # teacher5 = Teacher("Stephane", "Bonnevay", 10, [slots[random.randint(0, len(slots)-1)] for i in range(random.randint(0, len(slots)-1))])

    # apm1 = Apm("Eric", "Judor", 11)
    # apm2 = Apm("Anne", "Hathaway", 12)
    # apm3 = Apm("Jonathan", "Coen", 13)
    # apm4 = Apm("Robert", "De Nirot", 14)

    # group1 = Group(teacher1, apm2, student1)
    # group2 = Group(teacher1, apm1, student2)
    # group3 = Group(teacher2, apm3, student3)
    # group4 = Group(teacher2, apm4, student4)
    # group5 = Group(teacher3, apm4, student5)

    # for slot in slots:
    #     sm.addSlot(slot)

    # sm.addGroup(group1)
    # sm.addGroup(group2)
    # sm.addGroup(group3)
    # sm.addGroup(group4)
    # sm.addGroup(group5)

    # sm.addTeacher(teacher1)
    # sm.addTeacher(teacher2)
    # sm.addTeacher(teacher3)
    # sm.addTeacher(teacher4)
    # sm.addTeacher(teacher5)

    import names

    nb_students = 24
    nb_teachers = 10
    nb_apms = 35

    students = list()
    for i in range(nb_students):
        students.append(Person(names.get_first_name(), names.get_last_name(), i))

    teachers = list()
    for i in range(nb_teachers):
        contraintes = set()
        for j in range(random.randint(0, len(PERIODS)-1)):
            contraintes.add(random.choice(PERIODS))
        teachers.append(Teacher(names.get_first_name(), names.get_last_name(), i+nb_students, list(contraintes)))

    apms = list()
    for i in range(nb_apms):
        apms.append(Apm(names.get_first_name(), names.get_last_name(), i+nb_students+nb_teachers))
    
    groups = list()
    for i in range(nb_students):
        groups.append(Group(teachers[random.randint(0, nb_teachers-1)], apms[random.randint(0, nb_apms-1)], students[i]))

    for period in PERIODS:
        sm.addPeriod(Period(period[0], period[1], len(CLASSES)))

    for i in range(nb_students):
        sm.addGroup(groups[i])

    for i in range(nb_teachers):
        sm.addTeacher(teachers[i])

    


    
    pop = Population(sm, 100, True)
    #pop.schedules[random.randint(0, 49)]._toString()
    #pop.getBestScore()._toString()

    ga = GA(sm)
    bestScores = list()
    years = list()
    pop = ga.evolvePopulation(pop)
    years.append(0)
    bestScores.append((pop.getBestScore().getScore()))
    for i in tqdm(range(0, 100)):
        pop = ga.evolvePopulation(pop)
        years.append(i+1)
        bestScores.append(pop.getBestScore().getScore())

    bestPop = pop.getBestScore()
    bestPop._toString()
    print(bestPop.isCorrect())

    
    import numpy as np
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()
    ax.plot(years, bestScores)

    plt.show()

    print(pop.getBestScore().isCorrect())
    print(pop.getBestScore().getErrors())
    writer = WriterExcel("test.xlsx")
    writer.writeSchedule(pop.getBestScore())
    writer.writeConstraintes(sm.teachers, pop.getBestScore())
